# 2
# 엑셀파일의 데이터를 읽어와 각 품목의 총 지출 금액을 구해서 결과를 딕셔너리로 만들어 리스트로 담아주세요.
# {"품목" : "바나나", "총지출" : 5000}
# {"품목" : "커피", "총지출" : 10000}
# {"품목" : "책", "총지출" : 15000}
# {"품목" : "펜", "총지출" : 8000}

from openpyxl import load_workbook
wb = load_workbook('test2.xlsx')
ws = wb.active


print(wb.sheetnames)

header = ["품목", "가격", "수량", "날짜"]

items = []


for row in ws.iter_rows(min_row=2) :
    item = { "품목" : row[0].value,"총지출" : row[1].value * row[2].value}
    items.append(item)
       
        
print(items)
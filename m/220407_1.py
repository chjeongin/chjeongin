# 복습
# 변수, 함수 < 클래스 < 모듈(파일) < 패키지(폴더)
# 모듈
# 같은 폴더 안에 있을 경우 이름으로 호출

from test_pkg.test_mod import print_test

print_test()

# 클래스 ==> 데이터 + 함수 묶어놓은 것.

class Person :
    # 이름, 나이, 고향
    
    def __init__(self, name, age, home) : #self --> 객체
        self.name = name
        self.age = age
        self.home = home
    def introduce(self)     :
        print("{}사는 {}살 {}입니다".format(self.home, self.age, self.name))
        
p1 = Person("춘식이",1, "서울") #객체 생성 --> __init__()이 실행됨
p2 = Person("라이언", 5, "서울")

print(p1.name)
print(p1.age)

p1.introduce()



print("-"*50)

from openpyxl import Workbook
wb1 = Workbook() # 워크북 객체 // wb1 => 엑셀 파일

# 시트만들기
wb1.create_sheet("my_sheet1")
wb1.create_sheet("my_sheet2")

ws1 = wb1.active # 가장 앞쪽 시트를 활성화
ws1.title = "New Title"
ws2 = wb1["my_sheet2"] # 이름이 my_sheet2인 시트 선택

# 파일저장
wb1.save("test1.xlsx")




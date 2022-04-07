
from openpyxl import Workbook
wb1 = Workbook() # 워크북 객체 // wb1 => 엑셀 파일

# 시트만들기
wb1.create_sheet("my_sheet1")


ws1 = wb1.active # 가장 앞쪽 시트를 활성화
ws1.title = "New Title"
ws1 = wb1["my_sheet1"] # 이름이 my_sheet1인 시트 선택

# 파일저장
wb1.save("test1.xlsx")
ws1["A1"] = "No"
ws1["A2"] = 1
ws1["B1"] = "name"
ws1["b2"] = "춘식이"
ws1["C1"] = "favorite"
ws1["c2"] = "고구마"
target_cell = ws1.cell(row = 2, column = 3)
print(target_cell.value)


cell_range1 = ws1['A1':'C2']
# 이중리스트 ==> 행렬을 표현할 때 사용.

for row in cell_range1 :
    for cell in row :
        print(cell.value)


for row in ws1.iter_rows(min_row= 1,max_row=2, min_col= 1, max_col=3) :
    for cell in row :
        print(cell)

for row in ws1.rows :
    for cell in row :
        cell.value = '귀욤 춘식이' # 데이터 수정
    

# 행단위 추가 
# 가장 마지막 행에 추가 ==> append
ws1.append(["번호", "제목", "내용"])
ws1.append([1, "춘식이 겟 아이템", "고구마"])
ws1.append([2, "춘식이 기분", "기쁨"])




# 빈행 추가
ws1.insert_rows(3)

# 한 행을 선택
print(ws1[3])
row = ws1[3]
row[0].value = "Favorite"

# 행삭제
ws1.delete_rows(1)


wb1.save("test1.xlsx")

# 판다스 ==> 엑셀 작업 수월하게 도와줌.




